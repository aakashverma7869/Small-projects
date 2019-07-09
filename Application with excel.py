import openpyxl as xl
fileName = "Book1.xlxs"
isOpen = False
heading = []
data = []
book = None
sheet = None

def choose_y_n(string):
    choice = input(string + " (y/n) : ").lower()
    while True:
        if choice == 'n':
            return 'n'
        elif choice == 'y':
            return 'y'
        else:
            choice = input('invalid choice .... select either (y/n) : ').lower()


def choose_options(string, options):
    option_string = '/'.join(options)
    string += '(' + option_string + ') : '
    choice = input(string).upper()
    while True:
        if choice in options:
            return choice
        else:
            choice = input('invalid choice.......please try again with ('+ option_string + ') : ').upper()


def createWorkbook():
	global isOpen
	global book
	global sheet
	global data
	global heading
	try:
		if isOpen == False:
			isOpen = True
			book = xl.Workbook()
			sheet = book.active
			while True : 
				head = input("Enter Heading (Exit for exit): ").strip()
				if head.lower().find("exit") > -1 :
					break
				heading.append(head)
			print("\nAdd Row with (,) Separation or press Exit for Exit")
			print("\n***************************************************")
			for head in heading:
				print(head,end="\t")
			print("\n***************************************************")
			while True:
				row = input("Enter your data...(enter 'exit' to Exit this.)").strip()
				if row.lower().find("exit") > -1:
					break
				data.append(row.split(","))
		else:
			print(fileName, 'already opened, First closing this file\n')
			choice = input("Do you want to save changes to "+fileName).lower()
			if choice.find("save") > -1 or choice.find("yes") > -1 or choice.find('y') > -1:
				saveWorkbook()
	except Exception as e:
		print("File already in use : ",e)

def openWorkbook():
    global fileName
    global isOpen
    global sheet
    global book
    if not isOpen:
        try:
            tempName = input("Enter file Name to be Open (enter 'exit' to Exit) : ")
            if tempName.lower().find("exit") > -1:
            	return -1;
            
            book = xl.load_workbook(tempName)
            sheet = book.active
            fileName = tempName
            isOpen = True

        except Exception as e:
            print("No file exist with ")
            openWorkbook()
    else:
    	print('\n', fileName, 'already opened....')
    	saveWorkbook()

def displayWorkbook():
	global isOpen
	global book
	global sheet
	global data
	global heading
	global fileName
	if isOpen:
		active_sheet = book.active
		rows = active_sheet.max_row
		columns = active_sheet.max_column

		print('', end='   ')
		for i in range(1, columns+1):
			print('  ', chr(64+i),'  ', end='')
		print()
		for i in range(1, rows+1):
			for j in range(1, columns+1):
				if j == 1:
					print('  ', i, '  ', end='')
				print('  ', active_sheet.cell(row = i, column = j).value, '  ', end='')
			print()
		print()
	else:
		print('No WorkBook Opened yet...\n')
		i = openWorkbook()
		if not i:
			displayWorkbook()


def editWorkbook():
	global isOpen
	global book
	global sheet
	global data
	global heading

	if choose_y_n('Do you want to see the file before editing..') == 'y':
		displayWorkbook()

	if isOpen:
		rows = sheet.max_row
		columns = sheet.max_column
		row_list = [str(i) for i in range(1, rows+1)]
		col_list = [chr(64+i) for i in range(1, columns+1)]
		while True:
			row = input('Enter the row of the cell you want to edit (1,2,...) : ').strip()
			if row in row_list:
				break
			elif choose_y_n('Invalid row entered. Would you like to try again...') == 'n':
				return
		while True:
			col = input('Enter the column of the cell you want to edit (A, B. ..) : ').strip().upper()
			if col in col_list:
				break
			elif choose_y_n('Invalid column entered. Would you like to try again...') == 'n':
				return
		cell = sheet[col+row]
		cell.value = input('Enter value ...').strip()
		print('\nTo save changes please save file from menu\n')


from os import remove
def deleteWorkbook():
	global fileName
	global isOpen
	if isOpen:
		if choose_y_n('are you sure you want to delete this excel file : ' + fileName)=='y':
			os.remove(fileName)
	else:
		print('No file Opened yet')
	
def saveWorkbook():
	global fileName
	global heading
	global data
	global isOpen
	global book
	global sheet
	if isOpen == True:
		print('Now saving current file...')
		sheet.append(heading)
		for row in data:
			sheet.append(row)
		fileName = input("Enter File Name : ")
		if fileName.find(".xlsx") == -1:
			fileName = fileName + ".xlsx"
		try:
			book.save(fileName)
		except Exception as e:
			print(e)
	else:
		print("No file opened")

def closeWorkbook():
	global fileName
	global isOpen
	global book
	global sheet

	if choose_y_n('are you sure you want to close the excel file ', fileName) == 'y':
		isOpen = False
		book = None
		sheet = None
		fileName = ''

def main():
	global fileName
	while True:
		print("1. Create new Workbook")
		print("2. Open Workbook")
		print("3. Display Workbook Data")
		print("4. Edit Workbook Data")
		print("5. Delete Workbook")
		print("6. Save Workbook")
		print('7. Close Workbook')
		print("0. Exit")
		choice = int(choose_options("Enter Choice ", ['1', '2', '3', '4', '5', '6', '7', '0']))
		if choice == 1:
			createWorkbook()
		elif choice == 2:
			openWorkbook()
		elif choice == 3:
			displayWorkbook()
		elif choice == 4:
			editWorkbook()
		elif choice == 5:
			deleteWorkbook()
		elif choice == 6:
			saveWorkbook()
		elif choice == 6:
			closeWorkbook()
		elif choice == 0:
			break
	print('\n\n Bye bye..........')

if __name__ == "__main__":
	main()